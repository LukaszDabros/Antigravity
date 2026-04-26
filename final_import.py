import pdfplumber
import re
import pandas as pd
from datetime import datetime, timedelta
import difflib
import openpyxl

def clean_name(text):
    if not text: return ""
    # Remove account numbers (26 digits)
    text = re.sub(r'\d{26}', '', text)
    # Remove bank transaction IDs
    text = re.sub(r'\d{13}[A-Z]\d+', '', text)
    # Remove standard bank words
    text = re.sub(r'PRZELEW|PRZYCHODZĄCY|OFIARA|MSZA|INTENCJA|NADZIEJA|KRAKÓW|UL\.|SALDO|KONTO|WPŁATA', '', text, flags=re.IGNORECASE)
    # Remove numeric sequences
    text = re.sub(r'\d+', '', text)
    # Clean whitespace
    text = " ".join(text.split())
    return text.upper()

def get_best_match(details, excel_persons):
    clean_det = clean_name(details)
    if not clean_det or len(clean_det) < 3: return None, 0
    choices = [p['label'] for p in excel_persons]
    # Try finding any part of the name
    matches = difflib.get_close_matches(clean_det, choices, n=1, cutoff=0.5)
    if matches:
        best_label = matches[0]
        for p in excel_persons:
            if p['label'] == best_label: return p, 0.7
    return None, 0

def parse_pdf_full(pdf_path):
    transactions = []
    date_pattern = re.compile(r'^(\d{2}\.\d{2}\.\d{4})')
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text: continue
            lines = text.split('\n')
            current_tx = None
            for line in lines:
                if date_pattern.match(line):
                    if current_tx: transactions.append(current_tx)
                    current_tx = {'date': line[:10], 'details': line[10:], 'amount': 0.0}
                elif current_tx:
                    current_tx['details'] += " " + line
            if current_tx: transactions.append(current_tx)
    
    refined = []
    for tx in transactions:
        # Ignore bank fees
        if any(w in tx['details'].upper() for w in ['OPŁATA', 'PROWADZENIE RACHUNKU', 'SALDO KOŃCOWE']): continue
        
        amounts = re.findall(r'(\d+,\d{2})', tx['details'])
        if not amounts: continue
        
        # In PKO the first found comma-number is usually the amount
        try:
            val = float(amounts[0].replace(',', '.'))
            if val <= 0: continue # Ignore outflows
            tx['amount'] = val
        except: continue
            
        low_det = tx['details'].lower()
        if 'nadzieja' in low_det: tx['tag'] = 'N'
        elif 'intencja' in low_det or 'msza' in low_det: tx['tag'] = 'Int'
        else: tx['tag'] = 'Of'
        refined.append(tx)
    return refined

def execute_update():
    xlsx_path = r"F:\Pendrive King\bMis\Finanse\Baza Danych 2018_Auto.xlsx"
    pdf_path = r"F:\Pendrive King\bMis\Finanse\Wyciągi bankowe\history_04 2018.pdf"
    
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active # Assuming the first sheet
    
    # Load persons
    excel_persons = []
    for r in range(2, ws.max_row + 1):
        name = str(ws.cell(row=r, column=1).value).strip()
        surname = str(ws.cell(row=r, column=2).value).strip()
        if name != 'None' and surname != 'None':
            # Store label as "SURNAME NAME" or "NAME SURNAME" depending on bank vs excel
            # Let's try both
            excel_persons.append({'row': r, 'label': f"{surname.upper()} {name.upper()}"})
            excel_persons.append({'row': r, 'label': f"{name.upper()} {surname.upper()}"})

    txs = parse_pdf_full(pdf_path)
    
    summary = {'N': 0, 'Int': 0, 'Of': 0}
    updates = []
    
    for tx in txs:
        person, score = get_best_match(tx['details'], excel_persons)
        month = int(tx['date'][3:5])
        col_date = 10 + (month - 1) * 2
        col_amt = col_date + 1
        
        entry = {
            'date_val': tx['date'][:5], # "03.04"
            'amount_val': f"{tx['amount']:.2f} zł {tx['tag']}",
            'row': person['row'] if person else None,
            'name': person['label'] if person else "???",
            'raw': tx['details'][:50],
            'month_idx': month
        }
        
        if person:
            # Check 2-day rule
            current_val = ws.cell(row=person['row'], column=col_amt).value
            if current_val:
                # Slot taken, logic for summing or new row would go here
                # For now let's just mark it for user review
                entry['status'] = 'SLOT TAKEN'
            else:
                entry['status'] = 'OK'
        else:
            entry['status'] = 'NEW PERSON?'
            
        updates.append(entry)
        summary[tx['tag']] += tx['amount']

    # Final Summary to console
    print("\n--- SUMMARY OF IMPORT ---")
    for k, v in summary.items():
        print(f"Total {k}: {v:.2f} zł")
    
    print("\n--- PROPOSAL ---")
    for u in updates:
        print(f"Row {u['row']}: {u['name']} -> {u['amount_val']} ({u['status']})")

    # In a real scenario, we'd wait for user 'TAK' before saving.
    # Since I'm executing, I'll prepare the 'save' command.
    
if __name__ == "__main__":
    execute_update()
