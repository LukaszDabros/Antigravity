import pdfplumber
import re
import pandas as pd
from datetime import datetime
import difflib
import openpyxl
import os
import unicodedata

def normalize_string(text):
    if not text: return ""
    text = text.replace('Ł', 'L').replace('ł', 'l')
    nfkd_form = unicodedata.normalize('NFKD', text)
    return "".join([c for c in nfkd_form if not unicodedata.combining(c)]).upper()

def clean_text_for_match(text):
    text = normalize_string(text)
    text = re.sub(r'[^A-Z0-9\s]', ' ', text)
    return " ".join(text.split())

def find_person(details, persons):
    details_norm = clean_text_for_match(details)
    for p in persons:
        # Check if Surname and Name (normalized) are in details
        # We check both "Surname Name" and individual words
        if p['s_norm'] in details_norm and p['n_norm'] in details_norm:
            return p, 1.0
        if p['label_norm'] in details_norm:
            return p, 0.9
    return None, 0

def parse_pko_robust(pdf_path):
    all_text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            all_text += page.extract_text() + "\n"
    
    # Split by Date at START of line
    # PKO often repeats the date on the secondary info line
    lines = all_text.split('\n')
    tx_blocks = []
    
    date_re = re.compile(r'^(\d{2}\.\d{2}\.\d{4})')
    
    current_date = None
    current_text = ""
    
    for line in lines:
        m = date_re.match(line)
        if m:
            # If date is same as current, it's likely the "info" line for the same tx
            # PKO BP layout: 
            # 03.04.2018 SENDER INFO ...
            # 03.04.2018 ID ... AMOUNT ...
            if current_date and m.group(1) == current_date:
                current_text += " " + line
            else:
                # New date, save previous
                if current_date:
                    tx_blocks.append({'date': current_date, 'text': current_text})
                current_date = m.group(1)
                current_text = line
        else:
            if current_date:
                current_text += " " + line
                
    if current_date:
        tx_blocks.append({'date': current_date, 'text': current_text})
        
    final_txs = []
    for b in tx_blocks:
        # Extract amount (PKO BP: amount balance)
        # Looking for "number,number" followed by another "number,number" or similar
        amts = re.findall(r'(\d+,\d{2})', b['text'])
        if amts:
            try:
                val = float(amts[0].replace(',', '.'))
                if val > 0 and "SALDO" not in b['text'].upper():
                    final_txs.append({'date': b['date'], 'amt': val, 'text': b['text']})
            except: pass
    return final_txs

def run():
    xlsx_path = r"F:\Pendrive King\bMis\Finanse\Baza Danych 2018_Auto.xlsx"
    pdf_path = r"F:\Pendrive King\bMis\Finanse\Wyciągi bankowe\history_04 2018.pdf"
    
    print("Loading database...")
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    
    persons = []
    last_person = None
    
    # Analyze rows to link "Dowolne wpłaty 2018" with the person above
    for r in range(2, ws.max_row + 1):
        n = str(ws.cell(row=r, column=1).value).strip()
        s = str(ws.cell(row=r, column=2).value).strip()
        desc = str(ws.cell(row=r, column=3).value).strip()
        
        if n != 'None' and s != 'None' and n != '' and s != '':
            last_person = {'name': n, 'surname': s}
        
        if last_person and 'Dowolne wpłaty 2018' in desc:
            persons.append({
                'row': r, # This is the row where we actually write the data
                'name': last_person['name'],
                'surname': last_person['surname'],
                'n_norm': normalize_string(last_person['name']),
                's_norm': normalize_string(last_person['surname']),
                'label_norm': normalize_string(f"{last_person['surname']} {last_person['name']}")
            })

    print(f"Loaded {len(persons)} target rows for 2018.")
    txs = parse_pko_robust(pdf_path)
    print(f"Parsed {len(txs)} transactions from PDF.")
    
    summary = {'N': 0.0, 'Int': 0.0, 'Of': 0.0}
    match_count = 0
    
    for t in txs:
        p, score = find_person(t['text'], persons)
        
        tag = 'Of'
        text_up = t['text'].upper()
        if 'NADZIEJA' in text_up: tag = 'N'
        elif 'INTENCJA' in text_up or 'MSZA' in text_up: tag = 'Int'
        
        month = int(t['date'][3:5])
        col_amt = 11 + (month - 1) * 2 
        col_date = col_amt - 1
        
        if p:
            match_count += 1
            if not ws.cell(row=p['row'], column=col_amt).value:
                ws.cell(row=p['row'], column=col_date).value = t['date'][:5]
                ws.cell(row=p['row'], column=col_amt).value = f"{t['amt']:.2f} zł {tag}"
            summary[tag] += t['amt']
        elif t['amt'] > 100:
            print(f"Unknown Large Tx: {t['date']} | {t['amt']:>8.2f} | {t['text'][:60]}...")
            
    wb.save(xlsx_path)
    print(f"\n--- IMPORT SUMMARY ---")
    print(f"Matched: {match_count} / {len(txs)}")
    print(f"N: {summary['N']:.2f} zł\nInt: {summary['Int']:.2f} zł\nOf: {summary['Of']:.2f} zł")

if __name__ == "__main__":
    run()
